"""
Ghi list[InvoiceInfo] vào một Excel Table đã có sẵn trong workbook do người
dùng chọn ở UI (Output Excel), theo ExcelMapping.

Quy ước (ADR-006, ADR-002, ADR-004):
- ExcelWriter chỉ là tầng Persistence/I-O. Không chứa nghiệp vụ OCR, Template
  Matching, Validation hay Normalize - toàn bộ đã hoàn tất trước khi dữ liệu
  tới đây (list[InvoiceInfo]).
- Input: list[InvoiceInfo] (đã parse xong), path workbook do người dùng chọn,
  ExcelMapping (đã được Mapper.load() sẵn từ bên ngoài - ExcelWriter KHÔNG tự
  đọc mapping.json, không phụ thuộc Mapper/MappingError, tránh import chéo).
- Output: ExcelWriteResult (dữ liệu thuần). ExcelWriter không tự ghi
  report.txt hay log - ReportWriter đảm nhiệm việc đó (ADR liên quan đã chốt
  trong phiên thảo luận excel_writer).
- Lỗi cấp toàn cục (không mở được file, không tìm thấy Table, không lưu
  được) -> raise, dừng toàn bộ thao tác ghi. Field InvoiceInfo = None KHÔNG
  phải lỗi - ghi nhận vào ExcelWriteResult.warnings (ADR-032/033), không
  raise, không dừng việc ghi các invoice còn lại. Cột mapping không khớp
  header thực tế của workbook -> ghi nhận vào ExcelWriteResult.errors, bỏ
  qua cột đó, vẫn tiếp tục ghi các cột hợp lệ khác.
"""

from __future__ import annotations
from copy import copy
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from core.models import ExcelMapping, ExcelWriteResult, InvoiceInfo, InvoiceWarning


class WorkbookNotFoundError(Exception):
    """Workbook path không tồn tại hoặc không mở được."""


class ExcelTableNotFoundError(Exception):
    """Không tìm thấy Excel Table có tên khớp ExcelMapping.table trong workbook."""


class WorkbookSaveError(Exception):
    """Không lưu được workbook (VD: file đang mở, không có quyền ghi)."""


class ExcelWriter:
    """Ghi list[InvoiceInfo] vào Excel Table đã có sẵn trong workbook."""

    def write(
        self,
        path: Path,
        invoices: list[InvoiceInfo],
        mapping: ExcelMapping,
    ) -> ExcelWriteResult:
        """
        Ghi toàn bộ invoices vào Excel Table (mapping.table), lưu workbook
        1 lần sau khi ghi xong toàn bộ (đối xứng ADR-008 ở cấp file).
        Raises
        ------
        WorkbookNotFoundError
            path không tồn tại hoặc workbook không mở được.
        ExcelTableNotFoundError
            Không tìm thấy Excel Table tên mapping.table trong workbook.
        WorkbookSaveError
            Lưu workbook thất bại.
        """
        workbook = self._open_workbook(path)
        worksheet, table = self._find_table(workbook, mapping.table)

        column_index = self._read_header(worksheet, table)
        column_index, missing_errors = self._resolve_columns(mapping, column_index)

        warnings: list[InvoiceWarning] = []
        written = 0

        has_total = self._is_total_row_present(table)
        start_row = self._find_start_row(worksheet, table)
        count = len(invoices)

        min_col, min_row, max_col, old_max_row = range_boundaries(table.ref)
        data_end_row = old_max_row - 1 if has_total else old_max_row

        available_empty_rows = max(0, data_end_row - start_row + 1)

        if count > available_empty_rows and has_total:
            rows_to_insert = count - available_empty_rows
            insert_at = data_end_row + 1
            worksheet.insert_rows(insert_at, amount=rows_to_insert)

        current_row = start_row
        sample_row = min_row + 1
        for invoice in invoices:
            self._write_row(
                worksheet,
                current_row,
                invoice,
                column_index,
                warnings,
                sample_row=sample_row,
            )
            current_row += 1
            written += 1

        new_max_row = max(
            old_max_row,
            current_row - 1 + (1 if has_total else 0),
        )
        self._expand_table_ref(table, new_max_row)
        self._autofit_columns(worksheet, table)
        self._save_workbook(workbook, path)

        return ExcelWriteResult(
            total=len(invoices),
            written=written,
            warnings=tuple(warnings),
            errors=tuple(missing_errors),
        )

    # ------------------------------------------------------------------
    # Workbook / Table
    # ------------------------------------------------------------------

    @staticmethod
    def _open_workbook(path: Path):
        try:
            return openpyxl.load_workbook(path)
        except FileNotFoundError as error:
            raise WorkbookNotFoundError(
                f"Không tìm thấy file Excel: '{path}'."
            ) from error
        except OSError as error:
            raise WorkbookNotFoundError(
                f"Không mở được file Excel '{path}': {error}"
            ) from error

    @staticmethod
    def _find_table(workbook, table_name: str):
        for worksheet in workbook.worksheets:
            if table_name in worksheet.tables:
                return worksheet, worksheet.tables[table_name]

        raise ExcelTableNotFoundError(
            f"Không tìm thấy Excel Table '{table_name}' trong workbook."
        )

    @staticmethod
    def _is_total_row_present(table) -> bool:
        """Kiểm tra xem Table có dòng Total hay không."""
        return bool(
            getattr(table, "totalsRowCount", 0)
            or getattr(table, "totalsRowShown", False)
        )

    @staticmethod
    def _find_start_row(worksheet, table) -> int:
        """Tìm dòng trống đầu tiên ngay sau dòng dữ liệu cuối cùng trong Table."""
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        has_total = ExcelWriter._is_total_row_present(table)
        data_end_row = max_row - 1 if has_total else max_row

        last_data_row = min_row
        for r in range(min_row + 1, data_end_row + 1):
            if any(
                worksheet.cell(row=r, column=c).value is not None
                for c in range(min_col, max_col + 1)
            ):
                last_data_row = r

        return last_data_row + 1

    @staticmethod
    def _read_header(worksheet, table) -> dict[str, int]:
        """Đọc dòng header của Table -> {tên cột: chỉ số cột}."""
        min_col, min_row, max_col, _ = range_boundaries(table.ref)
        header: dict[str, int] = {}
        for col in range(min_col, max_col + 1):
            value = worksheet.cell(row=min_row, column=col).value
            if value is not None:
                header[str(value)] = col
        return header

    @staticmethod
    def _resolve_columns(
        mapping: ExcelMapping,
        header: dict[str, int],
    ) -> tuple[dict[int, str], list[str]]:
        """
        Đối chiếu mapping.columns với header thực tế của Table.
        Trả về {chỉ số cột: field_name} cho cột hợp lệ, và danh sách lỗi cho
        cột khai báo trong mapping nhưng không tồn tại trong workbook.
        """
        resolved: dict[int, str] = {}
        errors: list[str] = []

        for column_name, field_name in mapping.columns.items():
            col = header.get(column_name)
            if col is None:
                errors.append(
                    f"Cột mapping '{column_name}' không tồn tại trong Table "
                    f"'{mapping.table}' của workbook - bỏ qua cột này."
                )
                continue
            resolved[col] = field_name

        return resolved, errors

    @staticmethod
    def _expand_table_ref(table, new_max_row: int) -> None:
        min_col, min_row, max_col, _ = range_boundaries(table.ref)
        table.ref = (
            f"{get_column_letter(min_col)}{min_row}:"
            f"{get_column_letter(max_col)}{new_max_row}"
        )

    @staticmethod
    def _save_workbook(workbook, path: Path) -> None:
        try:
            workbook.save(path)
        except OSError as error:
            raise WorkbookSaveError(
                f"Không lưu được workbook '{path}': {error}"
            ) from error

    @staticmethod
    def _autofit_columns(
        worksheet, table, padding: int = 3, min_width: int = 12
    ) -> None:
        """Tự động điều chỉnh độ rộng các cột trong Table vừa vặn với nội dung."""
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        for col in range(min_col, max_col + 1):
            max_len = 0
            for row in range(min_row, max_row + 1):
                val = worksheet.cell(row=row, column=col).value
                if val is not None:
                    lines = str(val).split("\n")
                    for line in lines:
                        max_len = max(max_len, len(line))
            col_letter = get_column_letter(col)
            worksheet.column_dimensions[col_letter].width = max(
                max_len + padding, min_width
            )

    @staticmethod
    def _copy_cell_style(source_cell, target_cell) -> None:
        """Sao chép toàn bộ định dạng (style, format) từ ô mẫu sang ô mới."""
        if source_cell.has_style:
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.font = copy(source_cell.font)
            target_cell.fill = copy(source_cell.fill)
            target_cell.border = copy(source_cell.border)
            target_cell.alignment = copy(source_cell.alignment)
            target_cell.protection = copy(source_cell.protection)

    @staticmethod
    def _write_row(
        worksheet,
        row: int,
        invoice: InvoiceInfo,
        column_index: dict[int, str],
        warnings: list[InvoiceWarning],
        sample_row: int,
    ) -> None:
        for col, field_name in column_index.items():
            value = getattr(invoice, field_name, None)
            target_cell = worksheet.cell(row=row, column=col, value=value)

            source_cell = worksheet.cell(row=sample_row, column=col)
            ExcelWriter._copy_cell_style(source_cell, target_cell)

            if value is None:
                warnings.append(
                    InvoiceWarning(
                        source_file=invoice.source_file,
                        field_name=field_name,
                    )
                )
